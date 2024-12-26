import openpyxl
import random

def generate_question(topic, difficulty):
    # Replace with your specific question generation logic
    # Here's a simple example for string manipulation:
    if topic == "string_manipulation":
        if difficulty == "easy":
            return f"Reverse a given string: '{random.choice(['hello', 'world', 'python'])}'"
        elif difficulty == "medium":
            return f"Check if a given string is a palindrome: '{random.choice(['racecar', 'level', 'madam'])}'"
        elif difficulty == "hard":
            return "Implement a function to find the longest common subsequence between two strings."

# ... similar logic for other topics like arrays, objects, algorithms, etc.

def export_to_sheet(questions, filename):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "JavaScript Coding Questions"

    sheet['A1'] = "Question"
    sheet['B1'] = "Difficulty"
    sheet['C1'] = "Topic"

    row_num = 2
    for question, difficulty, topic in questions:
        sheet.cell(row=row_num, column=1).value = question
        sheet.cell(row=row_num, column=2).value = difficulty
        sheet.cell(row=row_num, column=3).value = topic
        row_num += 1

    workbook.save(filename)

# Example usage:
questions = []
topics = ["string_manipulation", "array_manipulation", "algorithms", "data_structures"]
difficulties = ["easy", "medium", "hard"]

for _ in range(100):  # Adjust the number as needed
    topic = random.choice(topics)
    difficulty = random.choice(difficulties)
    question = generate_question(topic, difficulty)
    questions.append((question, difficulty, topic))

export_to_sheet(questions, "javascript_coding_questions.xlsx")